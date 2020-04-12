
from openpyxl import Workbook
from openpyxl import load_workbook

# wb = load_workbook(filename = 'data/vendor-wise-record.xlsx')
# sheet = wb['Sheet1']
# print(sheet.max_row)
# sheet.delete_rows(idx=2, amount=1)

# noOfVendors = sheet.max_row
# print(noOfVendors)

# wb.save(filename = 'data/vendors.xlsx')
import os
from shutil import copyfile
import copy
# print(os.path.isfile(''))

# wb = load_workbook("data/month-wise-record.xlsx")
# sheet = wb['Sheet2']
# sheet["A1"] = "It works"
# wb.save("data/month-wise-record.xlsx")


# months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sept","Oct","Nov","Dec"]

# if not os.path.isfile("records/monthwise_record/"+str(1)+".xlsx"):
#     copyfile("data/month-wise-record.xlsx","records/monthwise_record/"+str(1)+".xlsx")
#     wb = load_workbook("records/monthwise_record/"+str(1)+".xlsx")
#     for i in months:
#         if i != "":
#             print(i)
#             print(wb[i].max_row)

#     wb.save("records/monthwise_record/"+str(1)+".xlsx")

# f = 1.52541254
# print(round(f,2))

# listOfBill = os.listdir('invoices/')
# for i in range(len(listOfBill)):
#     listOfBill[i] =listOfBill[i][]

# print(listOfBill)

# wb = load_workbook("test/2020.xlsx")
# sheet = wb['Apr']
# date = list(map(lambda x: int(x),sheet['I3'].value.split("/")))
# print(date)

# num = 0
# for i in range(12,27):
#     print(sheet['B'+str(i)].value)
#     if sheet['B'+str(i)].value:
#         num += 1
#     else:
#         break

# print(num)

# for xl_filename in os.listdir("test/productwise_record"):
#     print(xl_filename)
#     wb = load_workbook("test/productwise_record/"+xl_filename)
#     sheet = wb["Sheet1"]
#     invoiceNumber = "4111"
#     num = sheet.max_row
#     row_to_delete = []
#     for i in range(num,0,-1):
#         if sheet['d'+str(i)].value == invoiceNumber:
#             sheet.delete_rows(i)
#     wb.save("test/productwise_record/"+xl_filename)


# os.remove("invoices/45215-Rishab Munot.xlsx")
# a=0

# def aa():
#     if a == 1:
#         print(a)
#         return
#     print(0)

# aa()


# Step 1: unzip the window_bill file (in any folder you want)
# Step 2: open window_bill there are 3 files of use.
#             a) dashboard.exe 
#             b) Invoicing.exe
#             c) folder - Invoicing System 
# Step 3: Copy Invoicing System folder in C:/ (Strictly!!! No change in file,folder names)
# Step 4: Shortcuts of above 2 exe files can be created in desktop