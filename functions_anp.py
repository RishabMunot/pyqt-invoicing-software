# from functions_anp import *

# Add below class
    # def clear_data(self):
    #     print('clear')
    #     f_clear_data(self)

    # def add_product(self):
    #     print('add')
    #     f_add_product(self)

    # def new_product(self):
    #     print('new')
    #     f_new_product(self)

    # def edit_product(self):
    #     print('edit')
    #     f_edit_product(self)

    # def find_product(self):
    #     print('find')
    #     f_find_product(self)

    # def delete_product(self):
    #     print('del')
    #     f_delete_product(self)

# below call of retranslateUI
#         readyProducts(self)
#         self.button_clear.clicked.connect(self.clear_data)
#         self.button_add.clicked.connect(self.add_product)
#         self.button_new.clicked.connect(self.new_product)
#         self.button_find.clicked.connect(self.find_product)
#         self.button_delete.clicked.connect(self.delete_product)
#         self.button_edit.clicked.connect(self.edit_product)


import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot,Qt
from openpyxl import Workbook
from openpyxl import load_workbook
from shutil import copyfile
from os import rename

wb = 0
sheet = 1
data = []
names =[]

init_name = ""

def f_clear_data(self):
    self.product_name.setText("")
    self.uom.setText("")
    self.hsn_code.setText("")
    self.cgst.setText("")
    self.sgst.setText("")

def f_add_product(self):
    global wb,sheet
    id = self.product_id.text()
    name = self.product_name.text()
    uom = self.uom.text() +" " + self.comboBox.currentText()
    cgst = self.cgst.text()
    sgst = self.sgst.text()
    hsn = self.hsn_code.text()
    
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Warning)

    msg.setText("Do you want to add new Product?")
    msg.setWindowTitle("New Product")
    msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)	
    
    retval = msg.exec_()
    
    if (retval == 1024):
            saveAt = str(sheet.max_row+1)
            sheet['A'+saveAt] = id
            sheet['B'+saveAt] = name
            sheet['C'+saveAt] = uom
            sheet['D'+saveAt] = hsn
            sheet['E'+saveAt] = cgst
            sheet['F'+saveAt] = sgst

            wb.save("C:/Invoicing System/"+'data/products.xlsx')
            self.label.setText("Successfully Added")
            self.label.setStyleSheet('color: green')
            self.button_new.setEnabled(True)
            self.product_id.setEnabled(False)
            setInputs(self,False)
            self.button_clear.setEnabled(False)
            self.button_add.setEnabled(False)
            self.button_find.setEnabled(False)
            self.button_edit.setEnabled(False)
            self.button_delete.setEnabled(False)
            readyProducts(self)
            from os.path import isfile
            if not isfile("C:/Invoicing System/"+"Mousami Foods"+"/records/productwise_record/"+name+".xlsx"):
                copyfile("C:/Invoicing System/"+'data/product-wise-record.xlsx',"C:/Invoicing System/"+"Mousami Foods"+"/records/productwise_record/"+name+".xlsx")
            if not isfile("C:/Invoicing System/"+"Mitesh Dhanji Shah"+"/records/productwise_record/"+name+".xlsx"):
                copyfile("C:/Invoicing System/"+'data/product-wise-record.xlsx',"C:/Invoicing System/"+"Mitesh Dhanji Shah"+"/records/productwise_record/"+name+".xlsx")


def f_delete_product(self):
    global sheet,wb

    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)

    msg.setText("Do you want to Delete Product?")
    msg.setWindowTitle("Delete Product")
    msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)	
    
    retval = msg.exec_()
    
    if (retval == 1024):
        row = int(self.product_id.text())-999
        sheet.delete_rows(idx=row, amount=1)
        for i in range(row,sheet.max_row+1):
            sheet["A"+str(i)] = str(i+999)
        wb.save("C:/Invoicing System/"+'data/products.xlsx')
        f_clear_data(self)
        self.label.setText("Product Deleted")
        self.label.setStyleSheet('color: red') 
        self.button_new.setEnabled(True)
        self.product_id.setEnabled(False)
        self.button_clear.setEnabled(False)
        self.button_add.setEnabled(False)
        self.button_find.setEnabled(True)
        self.button_edit.setEnabled(False)
        self.button_delete.setEnabled(False)
        readyProducts(self)
        setInputs(self,False)

def f_new_product(self):
    self.label.setText("Add New Product")
    self.label.setStyleSheet('color: black')
    self.button_new.setEnabled(False)
    self.button_clear.setEnabled(True)
    self.button_add.setEnabled(True)
    self.button_find.setEnabled(True)
    self.button_edit.setEnabled(False)
    self.button_delete.setEnabled(False)
    f_clear_data(self)
    setInputs(self,True)
    readyProducts(self)

def f_edit_product(self):
    id = self.product_id.text()
    name = self.product_name.text()
    uom = self.uom.text() +" " + self.comboBox.currentText()
    cgst = self.cgst.text()
    sgst = self.sgst.text()
    hsn = self.hsn_code.text()

    msg = QMessageBox()
    msg.setIcon(QMessageBox.Warning)

    msg.setText("Do you want to Edit Product?")
    msg.setWindowTitle("Edit Product")
    msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)	
    
    retval = msg.exec_()
    
    if (retval == 1024):
        if init_name == name:
            pass
        else:
            self.product_name.setText(name)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Do you want to change Product Name?")
            msg.setWindowTitle("Edit Product")
            msg.setStandardButtons(QMessageBox.Yes|QMessageBox.No)	
            retval = msg.exec_()
            if retval == 16384:
                rename("C:/Invoicing System/Mousami Foods/records/productwise_record/"+init_name+".xlsx","C:/Invoicing System/Mousami Foods/records/productwise_record/"+name+".xlsx")
                rename("C:/Invoicing System/Mitesh Dhanji Shah/records/productwise_record/"+init_name+".xlsx","C:/Invoicing System/Mitesh Dhanji Shah/records/productwise_record/"+name+".xlsx")
               
            
        saveAt = str(int(self.product_id.text())-999)
        sheet['A'+saveAt] = id
        sheet['B'+saveAt] = name
        sheet['C'+saveAt] = uom
        sheet['D'+saveAt] = hsn
        sheet['E'+saveAt] = cgst
        sheet['F'+saveAt] = sgst

        wb.save("C:/Invoicing System/"+'data/products.xlsx')
        self.label.setText("Successfully Edited")
        self.label.setStyleSheet('color: green')
        self.button_new.setEnabled(True)
        self.product_id.setEnabled(False)
        setInputs(self,False)
        self.button_clear.setEnabled(False)
        self.button_add.setEnabled(False)
        self.button_find.setEnabled(False)
        self.button_edit.setEnabled(False)
        self.button_delete.setEnabled(False)
        readyProducts(self)

def f_find_product(self):
    global data,names,init_name
    aa = self.listOfProducts.currentItem().text().split("-")
    init_name = aa[0]
    pro_name = aa[0]
    pro_uom = aa[1]
    find = 0
    for i in range(len(data)):
        if data[i][1] == pro_name and data[i][2] == pro_uom:
            find = i
    print(find)
    setInputs(self,True)
    self.button_edit.setEnabled(True)
    self.button_delete.setEnabled(True)
    self.button_new.setEnabled(False)
    self.button_clear.setEnabled(False)
    self.button_new.setEnabled(False)
    self.product_id.setText(data[find][0])
    self.product_name.setText(data[find][1])
    self.uom.setText(data[find][2].split(" ")[0])
    index = self.comboBox.findText(data[find][2].split(" ")[1], Qt.MatchFixedString)
    if index >= 0:
         self.comboBox.setCurrentIndex(index)

    self.hsn_code.setText(data[find][3])
    self.cgst.setText(data[find][4])
    self.sgst.setText(data[find][5])

def readyProducts(self):

    global wb,sheet,data,names
    wb = load_workbook("C:/Invoicing System/"+'data/products.xlsx')
    sheet = wb['Sheet1']
    names = []
    data = []
    for i in range(2,sheet.max_row+1):
        name = sheet['B'+str(i)].value
        id = sheet['A'+str(i)].value
        uom = sheet['C'+str(i)].value
        hsn = sheet['D'+str(i)].value
        cgst = sheet['E'+str(i)].value
        sgst = sheet['F'+str(i)].value
        data.append([id,name,uom,hsn,cgst,sgst])
        la = name + "-" + uom
        names.append(la)
    print(data,names)
    self.listOfProducts.clear()
    self.listOfProducts.addItems(names)
    self.product_id.setText(str(1000 + sheet.max_row))

def setInputs(self,aa):
    self.product_name.setEnabled(aa)
    self.uom.setEnabled(aa)
    self.comboBox.setEnabled(aa)
    self.hsn_code.setEnabled(aa)
    self.cgst.setEnabled(aa)
    self.sgst.setEnabled(aa)