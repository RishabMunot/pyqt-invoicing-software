# from functions_anv import *

# Add this below class

    # def clear_data(self):
    #     print('clear')
    #     f_clear_data(self)
    # def add_vendor(self):
    #     print('add')
    #     f_add_vendor(self)
    # def new_vendor(self):
    #     print('new')
    #     f_new_vendor(self)

import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot
from openpyxl import Workbook
from openpyxl import load_workbook

def f_clear_data(self):
    self.vendor_name.setText("")
    self.add_line1.setText("")
    self.add_line2.setText("")
    self.add_city.setText("")
    self.add_pincode.setText("")
    self.add_district.setText("")
    self.add_state.setText("")
    self.vedor_gst_no.setText("")
    self.vendor_pan_no.setText("")
    self.vendor_name.setText("")

def f_add_vendor(self):
    name = self.vendor_name.text()
    line1 = self.add_line1.text()
    line2 = self.add_line2.text()
    city = self.add_city.text()
    pincode = self.add_pincode.text()
    dist = self.add_district.text()
    state = self.add_state.text()
    gstNO= self.vedor_gst_no.text()
    panNo = self.vendor_pan_no.text()
    

    msg = QMessageBox()
    msg.setIcon(QMessageBox.Warning)

    msg.setText("Do you want to add new Vendor?")
    msg.setWindowTitle("New Vendor")
    msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)	
    
    retval = msg.exec_()
    
    if (retval == 1024):
            wb = load_workbook(filename = 'data/vendors.xlsx')
            sheet = wb['Sheet1']
            saveAt = str(sheet.max_row+1)
            sheet['A'+saveAt] = name
            sheet['B'+saveAt] = line1
            sheet['C'+saveAt] = line2
            sheet['D'+saveAt] = city
            sheet['E'+saveAt] = dist
            sheet['F'+saveAt] = state
            sheet['G'+saveAt] = pincode
            sheet['H'+saveAt] = gstNO
            sheet['I'+saveAt] = panNo

            wb.save('data/vendors.xlsx')

            self.label.setText("Successfully Added")
            self.label.setStyleSheet('color: green')
            self.button_new.setEnabled(True)
            self.vendor_name.setEnabled(False)
            self.add_line1.setEnabled(False)
            self.add_line2.setEnabled(False)
            self.add_city.setEnabled(False)
            self.add_pincode.setEnabled(False)
            self.add_district.setEnabled(False)
            self.add_state.setEnabled(False)
            self.vedor_gst_no.setEnabled(False)
            self.vendor_pan_no.setEnabled(False)
            self.vendor_name.setEnabled(False)
            self.button_clear.setEnabled(False)
            self.button_add.setEnabled(False)

    
def f_new_vendor(self):
    f_clear_data(self)
    self.label.setText("Add New Vendor")
    self.button_new.setEnabled(False)
    self.label.setStyleSheet('color: black')
    self.vendor_name.setEnabled(True)
    self.add_line1.setEnabled(True)
    self.add_line2.setEnabled(True)
    self.add_city.setEnabled(True)
    self.add_pincode.setEnabled(True)
    self.add_district.setEnabled(True)
    self.add_state.setEnabled(True)
    self.vedor_gst_no.setEnabled(True)
    self.vendor_pan_no.setEnabled(True)
    self.vendor_name.setEnabled(True)
    self.button_clear.setEnabled(True)
    self.button_add.setEnabled(True)
