
# add this below call of retranslate

        # from functions_bill import setForm
        # setForm(self)

from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox,QCompleter
from PyQt5.QtGui import QIcon,QValidator
from PyQt5.QtCore import pyqtSlot,QDate,QEvent,QObject
from openpyxl import Workbook
from openpyxl import load_workbook
from shutil import copyfile
from amountInWords import convert
import os
from openpyxl.styles.borders import Border, Side

#data
products = {}
product_names = []
vendors = {}
vendors_names = []
months = ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sept","Oct","Nov","Dec"]

#inputs
el_productnames = []
products_inputs = []
vendor_inputs = []
final_outputs = []

listOfBill = []
productData = []

x = ""
previousBillNumber= 0
previousVendorName= 0
thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))

products_ips = []

def setForm(self):
    try:
        global x
        x = self
        self.date_edit.setDate(QDate.currentDate())
        import_data()
        group_inputs(self)
        set_functionalities(self)
        import_bill_data()
        self.listOfBills.clear()
        self.listOfBills.addItems(listOfBill)
        self.print_bill.setEnabled(False)
        self.edit_save.setEnabled(False)
    except Exception:
        pass

def import_data():
    try:
        global products,product_names,vendors,vendors_names
        wb = load_workbook('data/products.xlsx')
        product_sheet = wb['Sheet1']
        for i in range(2,product_sheet.max_row+1):
            name = product_sheet['B'+str(i)].value
            id = product_sheet['A'+str(i)].value
            uom = product_sheet['C'+str(i)].value
            hsn = product_sheet['D'+str(i)].value
            cgst = product_sheet['E'+str(i)].value
            sgst = product_sheet['F'+str(i)].value
            products[name+"-"+uom] = [id,name,hsn,uom,cgst,sgst]
            product_names.append(name+"-"+uom)
        
        wb = load_workbook('data/vendors.xlsx')
        product_sheet = wb['Sheet1']
        for i in range(2,product_sheet.max_row+1):
            name = product_sheet['A'+str(i)].value
            line1 = product_sheet['B'+str(i)].value
            line2 = product_sheet['C'+str(i)].value
            city = product_sheet['D'+str(i)].value
            dist = product_sheet['E'+str(i)].value
            state = product_sheet['F'+str(i)].value
            pincode = product_sheet['G'+str(i)].value
            gstNO = product_sheet['H'+str(i)].value
            panNo = product_sheet['I'+str(i)].value
            vendors[name] = [name,line1,line2,city,dist,state,pincode,gstNO,panNo]
            vendors_names.append(name)
    except Exception:
        pass

def import_bill_data():
    try:
        global listOfBill
        listOfBill = os.listdir('invoices/')
        for i in range(len(listOfBill)):
            listOfBill[i] =listOfBill[i][:-5]
    except Exception:
        pass

def group_inputs(self):
    global el_productnames,products_inputs,vendor_inputs,final_outputs
    el_productnames.append(self.p_desc_1)
    el_productnames.append(self.p_desc_2)
    el_productnames.append(self.p_desc_3)
    el_productnames.append(self.p_desc_4)
    el_productnames.append(self.p_desc_5)
    el_productnames.append(self.p_desc_6)
    el_productnames.append(self.p_desc_7)
    el_productnames.append(self.p_desc_8)
    el_productnames.append(self.p_desc_9)
    el_productnames.append(self.p_desc_10)
    el_productnames.append(self.p_desc_11)
    el_productnames.append(self.p_desc_12)
    el_productnames.append(self.p_desc_13)
    el_productnames.append(self.p_desc_14)
    el_productnames.append(self.p_desc_15)

    products_inputs.append([self.p_desc_1,self.p_hsn_1,self.p_uom_1,self.p_qty_1,self.p_rate_1,self.p_taxval_1,
    self.p_gst_per_1,self.p_gst_amt_1,self.p_sgst_per_1,self.p_sgst_amt_1,self.p_total_1])
    
    products_inputs.append([self.p_desc_2,self.p_hsn_2,self.p_uom_2,self.p_qty_2,self.p_rate_2,self.p_taxval_2,
    self.p_gst_per_2,self.p_gst_amt_2,self.p_sgst_per_2,self.p_sgst_amt_2,self.p_total_2])

    products_inputs.append([self.p_desc_3,self.p_hsn_3,self.p_uom_3,self.p_qty_3,self.p_rate_3,self.p_taxval_3,
    self.p_gst_per_3,self.p_gst_amt_3,self.p_sgst_per_3,self.p_sgst_amt_3,self.p_total_3])
    
    products_inputs.append([self.p_desc_4,self.p_hsn_4,self.p_uom_4,self.p_qty_4,self.p_rate_4,self.p_taxval_4,
    self.p_gst_per_4,self.p_gst_amt_4,self.p_sgst_per_4,self.p_sgst_amt_4,self.p_total_4])

    products_inputs.append([self.p_desc_5,self.p_hsn_5,self.p_uom_5,self.p_qty_5,self.p_rate_5,self.p_taxval_5,
    self.p_gst_per_5,self.p_gst_amt_5,self.p_sgst_per_5,self.p_sgst_amt_5,self.p_total_5])
    
    products_inputs.append([self.p_desc_6,self.p_hsn_6,self.p_uom_6,self.p_qty_6,self.p_rate_6,self.p_taxval_6,
    self.p_gst_per_6,self.p_gst_amt_6,self.p_sgst_per_6,self.p_sgst_amt_6,self.p_total_6])

    products_inputs.append([self.p_desc_7,self.p_hsn_7,self.p_uom_7,self.p_qty_7,self.p_rate_7,self.p_taxval_7,
    self.p_gst_per_7,self.p_gst_amt_7,self.p_sgst_per_7,self.p_sgst_amt_7,self.p_total_7])
    
    products_inputs.append([self.p_desc_8,self.p_hsn_8,self.p_uom_8,self.p_qty_8,self.p_rate_8,self.p_taxval_8,
    self.p_gst_per_8,self.p_gst_amt_8,self.p_sgst_per_8,self.p_sgst_amt_8,self.p_total_8])

    products_inputs.append([self.p_desc_9,self.p_hsn_9,self.p_uom_9,self.p_qty_9,self.p_rate_9,self.p_taxval_9,
    self.p_gst_per_9,self.p_gst_amt_9,self.p_sgst_per_9,self.p_sgst_amt_9,self.p_total_9])
    
    products_inputs.append([self.p_desc_10,self.p_hsn_10,self.p_uom_10,self.p_qty_10,self.p_rate_10,self.p_taxval_10,
    self.p_gst_per_10,self.p_gst_amt_10,self.p_sgst_per_10,self.p_sgst_amt_10,self.p_total_10])

    products_inputs.append([self.p_desc_11,self.p_hsn_11,self.p_uom_11,self.p_qty_11,self.p_rate_11,self.p_taxval_11,
    self.p_gst_per_11,self.p_gst_amt_11,self.p_sgst_per_11,self.p_sgst_amt_11,self.p_total_11])
    
    products_inputs.append([self.p_desc_12,self.p_hsn_12,self.p_uom_12,self.p_qty_12,self.p_rate_12,self.p_taxval_12,
    self.p_gst_per_12,self.p_gst_amt_12,self.p_sgst_per_12,self.p_sgst_amt_12,self.p_total_12])
    
    products_inputs.append([self.p_desc_13,self.p_hsn_13,self.p_uom_13,self.p_qty_13,self.p_rate_13,self.p_taxval_13,
    self.p_gst_per_13,self.p_gst_amt_13,self.p_sgst_per_13,self.p_sgst_amt_13,self.p_total_13])

    products_inputs.append([self.p_desc_14,self.p_hsn_14,self.p_uom_14,self.p_qty_14,self.p_rate_14,self.p_taxval_14,
    self.p_gst_per_14,self.p_gst_amt_14,self.p_sgst_per_14,self.p_sgst_amt_14,self.p_total_14])
    
    products_inputs.append([self.p_desc_15,self.p_hsn_15,self.p_uom_15,self.p_qty_15,self.p_rate_15,self.p_taxval_15,
    self.p_gst_per_15,self.p_gst_amt_15,self.p_sgst_per_15,self.p_sgst_amt_15,self.p_total_15])
    
    vendor_inputs = [self.vendor_name,self.add_line1,self.add_line2,self.add_city,self.add_district,self.add_state,self.add_pincode,self.vedor_gst_no,self.vendor_pan_no]

    final_outputs = [self.total_qty,self.total_tax_val,self.total_cgst_amt,self.total_sgst_amt, self.total_total,self.amt_bt,self.amt_cgst,self.amt_sgst,self.amt_gst,self.amt_at]

def set_functionalities(self):
    try:
    #adding completers
        completer = QCompleter(product_names)
        completer.setCaseSensitivity(False)
        for i in range(15):
            el_productnames[i].setCompleter(completer)


    #Listner on product name
        el_productnames[0].editingFinished.connect(lambda:fillDataProduct(1))
        el_productnames[1].editingFinished.connect(lambda:fillDataProduct(2))
        el_productnames[2].editingFinished.connect(lambda:fillDataProduct(3))
        el_productnames[3].editingFinished.connect(lambda:fillDataProduct(4))
        el_productnames[4].editingFinished.connect(lambda:fillDataProduct(5))
        el_productnames[5].editingFinished.connect(lambda:fillDataProduct(6))
        el_productnames[6].editingFinished.connect(lambda:fillDataProduct(7))
        el_productnames[7].editingFinished.connect(lambda:fillDataProduct(8))
        el_productnames[8].editingFinished.connect(lambda:fillDataProduct(9))
        el_productnames[9].editingFinished.connect(lambda:fillDataProduct(10))
        el_productnames[10].editingFinished.connect(lambda:fillDataProduct(11))
        el_productnames[11].editingFinished.connect(lambda:fillDataProduct(12))
        el_productnames[12].editingFinished.connect(lambda:fillDataProduct(13))
        el_productnames[13].editingFinished.connect(lambda:fillDataProduct(14))
        el_productnames[14].editingFinished.connect(lambda:fillDataProduct(15))

    #Listner on rate
        products_inputs[0][4].editingFinished.connect(lambda:fillValues(1))
        products_inputs[1][4].editingFinished.connect(lambda:fillValues(2))
        products_inputs[2][4].editingFinished.connect(lambda:fillValues(3))
        products_inputs[3][4].editingFinished.connect(lambda:fillValues(4))
        products_inputs[4][4].editingFinished.connect(lambda:fillValues(5))
        products_inputs[5][4].editingFinished.connect(lambda:fillValues(6))
        products_inputs[6][4].editingFinished.connect(lambda:fillValues(7))
        products_inputs[7][4].editingFinished.connect(lambda:fillValues(8))
        products_inputs[8][4].editingFinished.connect(lambda:fillValues(9))
        products_inputs[9][4].editingFinished.connect(lambda:fillValues(10))
        products_inputs[10][4].editingFinished.connect(lambda:fillValues(11))
        products_inputs[11][4].editingFinished.connect(lambda:fillValues(12))
        products_inputs[12][4].editingFinished.connect(lambda:fillValues(13))
        products_inputs[13][4].editingFinished.connect(lambda:fillValues(14))
        products_inputs[14][4].editingFinished.connect(lambda:fillValues(15))

    #Listner on qty
        products_inputs[0][3].editingFinished.connect(lambda:fillValues(1))
        products_inputs[1][3].editingFinished.connect(lambda:fillValues(2))
        products_inputs[2][3].editingFinished.connect(lambda:fillValues(3))
        products_inputs[3][3].editingFinished.connect(lambda:fillValues(4))
        products_inputs[4][3].editingFinished.connect(lambda:fillValues(5))
        products_inputs[5][3].editingFinished.connect(lambda:fillValues(6))
        products_inputs[6][3].editingFinished.connect(lambda:fillValues(7))
        products_inputs[7][3].editingFinished.connect(lambda:fillValues(8))
        products_inputs[8][3].editingFinished.connect(lambda:fillValues(9))
        products_inputs[9][3].editingFinished.connect(lambda:fillValues(10))
        products_inputs[10][3].editingFinished.connect(lambda:fillValues(11))
        products_inputs[11][3].editingFinished.connect(lambda:fillValues(12))
        products_inputs[12][3].editingFinished.connect(lambda:fillValues(13))
        products_inputs[13][3].editingFinished.connect(lambda:fillValues(14))
        products_inputs[14][3].editingFinished.connect(lambda:fillValues(15))

    #Listner on CSGT
        products_inputs[0][6].editingFinished.connect(lambda:fillValues(1))
        products_inputs[1][6].editingFinished.connect(lambda:fillValues(2))
        products_inputs[2][6].editingFinished.connect(lambda:fillValues(3))
        products_inputs[3][6].editingFinished.connect(lambda:fillValues(4))
        products_inputs[4][6].editingFinished.connect(lambda:fillValues(5))
        products_inputs[5][6].editingFinished.connect(lambda:fillValues(6))
        products_inputs[6][6].editingFinished.connect(lambda:fillValues(7))
        products_inputs[7][6].editingFinished.connect(lambda:fillValues(8))
        products_inputs[8][6].editingFinished.connect(lambda:fillValues(9))
        products_inputs[9][6].editingFinished.connect(lambda:fillValues(10))
        products_inputs[10][6].editingFinished.connect(lambda:fillValues(11))
        products_inputs[11][6].editingFinished.connect(lambda:fillValues(12))
        products_inputs[12][6].editingFinished.connect(lambda:fillValues(13))
        products_inputs[13][6].editingFinished.connect(lambda:fillValues(14))
        products_inputs[14][6].editingFinished.connect(lambda:fillValues(15))

    #Listner on SGST
        products_inputs[0][8].editingFinished.connect(lambda:fillValues(1))
        products_inputs[1][8].editingFinished.connect(lambda:fillValues(2))
        products_inputs[2][8].editingFinished.connect(lambda:fillValues(3))
        products_inputs[3][8].editingFinished.connect(lambda:fillValues(4))
        products_inputs[4][8].editingFinished.connect(lambda:fillValues(5))
        products_inputs[5][8].editingFinished.connect(lambda:fillValues(6))
        products_inputs[6][8].editingFinished.connect(lambda:fillValues(7))
        products_inputs[7][8].editingFinished.connect(lambda:fillValues(8))
        products_inputs[8][8].editingFinished.connect(lambda:fillValues(9))
        products_inputs[9][8].editingFinished.connect(lambda:fillValues(10))
        products_inputs[10][8].editingFinished.connect(lambda:fillValues(11))
        products_inputs[11][8].editingFinished.connect(lambda:fillValues(12))
        products_inputs[12][8].editingFinished.connect(lambda:fillValues(13))
        products_inputs[13][8].editingFinished.connect(lambda:fillValues(14))
        products_inputs[14][8].editingFinished.connect(lambda:fillValues(15))

        vendor_inputs[0].editingFinished.connect(fillVendor)

        completer = QCompleter(vendors_names)
        completer.setCaseSensitivity(False)
        self.vendor_name.setCompleter(completer)

    #Listener on Buttons
        self.edit_find.clicked.connect(findClicked)
        self.new_bill.clicked.connect(newClicked)
        self.edit_save.clicked.connect(editSaveCliked)
        self.save_bill.clicked.connect(saveBillClicked)
        self.print_bill.clicked.connect(printBill)
    except Exception:
        pass

def generateWarningForInt(x):
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("Input can only be a number")
    msg.setWindowTitle("Please enter number")
    msg.setStandardButtons(QMessageBox.Ok)
    retval = msg.exec_()
       
def fillDataProduct(x):
    try:
        if products_inputs[x-1][0].text() == "":
            products_inputs[x-1][3].setText("0")
            products_inputs[x-1][4].setText("0")
            products_inputs[x-1][2].setText("")
            products_inputs[x-1][1].setText("")
            fillValues(x)
            return

        pro = products.get(el_productnames[x-1].text())
        ip = products_inputs[x-1]
        if pro:
            ip[0].setText(pro[1]+"-" + pro[3])
            ip[1].setText(pro[2])
            ip[2].setText(pro[3])
            ip[6].setText(pro[4])
            ip[8].setText(pro[5])
    except Exception:
        pass

def fillVendor():
    try:
        ven = vendors.get(vendor_inputs[0].text())
        if ven:
            for i in range(9):
                vendor_inputs[i].setText(ven[i])
    except Exception:
        pass

def fillValues(a):
        
    total = 0

    try:
        qty = products_inputs[a-1][3].text()
        rate = products_inputs[a-1][4].text()
        if(qty== ""):
            products_inputs[a-1][3].setText("0")
        elif(rate == ""):
            products_inputs[a-1][4].setText("0")

        else:
            qty = int(qty)
            rate = int(rate)

        taxval = str(qty*rate)
        products_inputs[a-1][5].setText(taxval)
    except Exception:
        generateWarningForInt(a)
        pass

    try:
        taxval = products_inputs[a-1][5].text()
        cgst = products_inputs[a-1][6].text()
        sgst = products_inputs[a-1][8].text()
        total = 0
        if(taxval== ""):
            products_inputs[a-1][5].setText("0")            
        elif(cgst == ""):
            products_inputs[a-1][6].setText("0") 
        elif sgst == "":
            products_inputs[a-1][8].setText("0") 
        else:
            taxval = float(taxval)
            cgst = float(cgst)
            sgst = float(sgst)
            
            total = taxval*(1+ cgst*0.01 + sgst*0.01)
            total = round(total,2)

        products_inputs[a-1][7].setText(str(round(cgst*taxval*0.01,2)))
        products_inputs[a-1][9].setText(str(round(sgst*taxval*0.01,2)))

        products_inputs[a-1][10].setText(str(total))
    except Exception:
        generateWarningForInt(x)
        pass


    fillFinalValues()

def fillFinalValues():
    tot_qty = 0
    tot_taxval = 0
    tot_sgst = 0
    tot_cgst = 0
    tot_total = 0
    try:
        for i in range(15):
            tot_qty += int(products_inputs[i][3].text())
            tot_taxval += round(float(products_inputs[i][5].text()),2)
            tot_sgst += round(float(products_inputs[i][9].text()),2)
            tot_cgst += round(float(products_inputs[i][7].text()),2)
            tot_total += round(float(products_inputs[i][10].text()),2)
        
        final_outputs[0].setText(str(round(tot_qty,2)))
        final_outputs[1].setText(str(round(tot_taxval,2)))
        final_outputs[2].setText(str(round(tot_cgst,2)))
        final_outputs[3].setText(str(round(tot_sgst,2)))
        final_outputs[4].setText(str(round(tot_total,2)))
        final_outputs[5].setText(str(round(tot_taxval,2)))
        final_outputs[6].setText(str(round(tot_cgst,2)))
        final_outputs[7].setText(str(round(tot_sgst,2)))
        final_outputs[8].setText(str(round(tot_cgst+tot_sgst,2)))
        final_outputs[9].setText(str(round(tot_total,2)))

        x.total_amont_in_words.setText("Amount in words: INR "+convert(float(x.amt_at.text()))+" only")

    except Exception:
        pass

def findClicked():
    try:
        global previousBillNumber,previousVendorName
        xlfile = x.listOfBills.currentItem().text()

        wb = load_workbook("invoices/"+xlfile+".xlsx")
        sheet = wb['Sheet1']
        previousBillNumber = sheet['I2'].value
        previousVendorName = sheet['f5'].value
        x.invoice_number.setText(sheet['I2'].value)
        date = list(map(lambda x: int(x),sheet['I3'].value.split("/")))
        x.date_edit.setDate(QDate(date[2],date[1],date[0]))     
        vendor_inputs[0].setText(sheet['f5'].value)
        fillVendor()

        num = 0
        for i in range(12,27):
            if sheet['B'+str(i)].value:
                num += 1
            else:
                break
        for i in range(num):
            ch = 'B'
            for j in range(11):
                products_inputs[i][j].setText(sheet[ch+str(i+12)].value)
                ch = chr(ord(ch)+1)
        
        x.total_qty.setText(sheet['e27'].value)
        x.total_tax_val.setText(sheet['g27'].value)
        x.total_cgst_amt.setText(sheet['i27'].value)
        x.total_sgst_amt.setText(sheet['k27'].value)
        x.total_total.setText(sheet['l27'].value)

        x.amt_bt.setText(str(sheet['j29'].value))
        x.amt_cgst.setText(str(sheet['j30'].value))
        x.amt_sgst.setText(str(sheet['j31'].value))
        x.amt_gst.setText(str(sheet['j32'].value))
        x.amt_at.setText(str(sheet['j33'].value))
        
        x.print_bill.setEnabled(True)
        x.edit_save.setEnabled(True)

    except Exception:
        pass

def editSaveCliked():
    try:
        invoiceNumber = x.invoice_number.text()
        ven_name = x.vendor_name.text()

        if invoiceNumber == "" or ven_name == "":
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Please enter invoice number and vendor name?")
            msg.setWindowTitle("Incomplete form")
            msg.setStandardButtons(QMessageBox.Ok)        
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Do you want to save the changes in the bill?")
            msg.setWindowTitle("Edit Bill")
            msg.setStandardButtons(QMessageBox.Ok|QMessageBox.Cancel)
            retval = msg.exec_()
            if retval == 1024:
                deleteExistingDataInRecords()
                storeBill(invoiceNumber,ven_name)
                
                x.label.setText("Bill Successfully Edited and saved")
                x.label.setStyleSheet('color: green')
    except Exception:
        pass

def newClicked():
    try:
        x.invoice_number.setText("")
        for a in vendor_inputs:
            a.setText("")

        for a in products_inputs:
            for b in a:
                b.setText("0")

        for a in products_inputs:
            for b in a[:2]:
                b.setText("")

        for a in final_outputs:
            a.setText("")
        disableViews(True)
        import_bill_data()
        x.listOfBills.clear()
        x.listOfBills.addItems(listOfBill)
        x.label.setText("Tax Invoice")
        x.label.setStyleSheet('color: black')
        x.save_bill.setEnabled(True)   
        x.edit_find.setEnabled(True)
        x.print_bill.setEnabled(False)   
        x.edit_save.setEnabled(False)
    except Exception:
        pass

def deleteExistingDataInRecords():

    date = x.date_edit.date()
    invoiceNumber = previousBillNumber
    
    #deleteing monthwise
    wb = load_workbook("records/monthwise_record/"+str(date.year())+".xlsx")
    sheet = wb[months[date.month()]]
    num = sheet.max_row
    for i in range(num,0,-1):
        if sheet['c'+str(i)].value == invoiceNumber:
            sheet.delete_rows(i)
    wb.save("records/monthwise_record/"+str(date.year())+".xlsx")

    #deleting vendorWise
    ven_name = x.vendor_name.text()
    wb = load_workbook("records/vendorwise_record/"+ven_name+".xlsx")
    sheet = wb['Sheet1']
    num = sheet.max_row
    for i in range(num,0,-1):
        if sheet['d'+str(i)].value == invoiceNumber:
            sheet.delete_rows(i)
    wb.save("records/vendorwise_record/"+ven_name+".xlsx")

    #deleting productwise
    for xl_filename in os.listdir("records/productwise_record"):
        wb = load_workbook("records/productwise_record/"+xl_filename)
        sheet = wb["Sheet1"]
        num = sheet.max_row
        row_to_delete = []
        for i in range(num,0,-1):
            if sheet['d'+str(i)].value == invoiceNumber:
                sheet.delete_rows(i)
        wb.save("records/productwise_record/"+xl_filename)

    os.remove("invoices/"+previousBillNumber+"-"+previousVendorName+".xlsx")

def saveBillClicked():
    
    invoiceNumber = x.invoice_number.text()
    ven_name = x.vendor_name.text()
    if previousBillNumber != invoiceNumber:
        if invoiceNumber == "" or ven_name == "":
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Please enter invoice number and vendor name?")
            msg.setWindowTitle("Incomplete form")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()        
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Do you want to save the bill?")
            msg.setWindowTitle("New Bill")
            msg.setStandardButtons(QMessageBox.Ok|QMessageBox.Cancel)
            retval = msg.exec_()
            if retval == 1024:
                storeBill(invoiceNumber,ven_name)
    else:
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("Please change invoice number")
        msg.setWindowTitle("Similar Invoice Number")
        msg.setStandardButtons(QMessageBox.Ok) 
        retval = msg.exec_()

def storeBill(invoiceNumber,ven_name):
    
    copyfile("data/bill_format.xlsx","invoices/"+invoiceNumber+"-"+ven_name+".xlsx")
    wb = load_workbook("invoices/"+invoiceNumber+"-"+ven_name+".xlsx")
    sheet = wb['Sheet1']
    
    sheet['I2'] = invoiceNumber
    date = x.date_edit.date()       
    sheet['I3'] = str(date.day())+"/"+str(date.month())+ "/" + str(date.year())
    sheet['f5'] = ven_name
    sheet['f6'] = vendor_inputs[1].text()
    sheet['f7'] = vendor_inputs[2].text()
    sheet['f8'] = vendor_inputs[3].text()+", "+vendor_inputs[4].text()+", "+vendor_inputs[5].text()+" - "+vendor_inputs[6].text()
    sheet['h9'] = vendor_inputs[7].text()

    num = 0
    for c in range(14,-1,-1):
        if el_productnames[c].text() == "":
            num += 1
        else:
            num = 15 - num
            break
    print(num)
    i=-1
    for q in range(num):
        if products_inputs[q][0].text() == "":
            continue
        i += 1
        ch = 'B'
        for j in range(11):
            sheet[ch+str(i+12)] = products_inputs[q][j].text()
            ch = chr(ord(ch)+1)
    
    sheet['e27'] = x.total_qty.text()
    sheet['g27'] = x.total_tax_val.text()
    sheet['i27'] = x.total_cgst_amt.text()
    sheet['k27'] = x.total_sgst_amt.text()
    sheet['l27'] = x.total_total.text()

    sheet['j29'] = float(x.amt_bt.text())
    sheet['j30'] = float(x.amt_cgst.text())
    sheet['j31'] = float(x.amt_sgst.text())
    sheet['j32'] = float(x.amt_gst.text())
    sheet['j33'] = float(x.amt_at.text())

    wb.save("invoices/"+invoiceNumber+"-"+ven_name+".xlsx")
    saveMonthWiseData()
    saveProductWise()
    saveVendorWise()
    
    import_bill_data()
    x.listOfBills.clear()
    x.listOfBills.addItems(listOfBill)
    x.label.setText("Bill Successfully saved")
    x.label.setStyleSheet('color: green')
    disableViews(False)
    x.new_bill.setEnabled(True)
    x.save_bill.setEnabled(False)  
    x.print_bill.setEnabled(False)  
    x.edit_find.setEnabled(False)
    x.edit_save.setEnabled(False)

def printBill():
    invoice_number = x.invoice_number.text()
    ven_name = vendor_inputs[0].text()
    os.startfile("invoices/"+invoice_number + "-" + ven_name+".xlsx",'print')

def disableViews(val):
    x.invoice_number.setEnabled(val)
    for a in vendor_inputs:
        a.setEnabled(val)
    for a in products_inputs:
        for b in a:
            b.setEnabled(val)

def saveMonthWiseData():
    numOfProducts = 0
    for c in range(14,-1,-1):
        if el_productnames[c].text() == "":
            numOfProducts += 1
        else:
            numOfProducts = 15 - numOfProducts
            break
    global productData
    productData = []
    date = x.date_edit.date()
    # product data is formed - 2d array with all the data filled in bill window
    for i in range(numOfProducts):
        unit = []
        if el_productnames[i].text() =="":
            continue

        for j in range(11):
            unit.append(products_inputs[i][j].text())
        productData.append(unit)

    ven_name = x.vendor_name.text()
    ven_gst = x.vedor_gst_no.text()
    invoice_number = x.invoice_number.text()

    if not os.path.isfile("records/monthwise_record/"+str(date.year())+".xlsx"):
        copyfile("data/month-wise-record.xlsx","records/monthwise_record/"+str(date.year())+".xlsx")

    wb = load_workbook("records/monthwise_record/"+str(date.year())+".xlsx")
    sheet = wb[months[date.month()]]
    numRowToStart = sheet.max_row +1
    print(numRowToStart)

    i = -1
    for q in range(numOfProducts):
        
        if el_productnames[q].text() == "":
            continue
        i+=1
        sheet["a"+str(numRowToStart+i)] = numRowToStart+i-2
        sheet["b"+str(numRowToStart+i)] = str(date.day())+"/"+str(date.month())+ "/" + str(date.year())        
        sheet["c"+str(numRowToStart+i)] = invoice_number
        sheet["d"+str(numRowToStart+i)] = ven_gst
        sheet["e"+str(numRowToStart+i)] = ven_name
        sheet["f"+str(numRowToStart+i)] = productData[i][0].split("-")[0]
        sheet["g"+str(numRowToStart+i)] = productData[i][1]
        sheet["h"+str(numRowToStart+i)] = productData[i][2]
        sheet["i"+str(numRowToStart+i)] = productData[i][3]
        sheet["j"+str(numRowToStart+i)] = productData[i][4]
        sheet["k"+str(numRowToStart+i)] = productData[i][5]
        sheet["l"+str(numRowToStart+i)] = productData[i][6]
        sheet["m"+str(numRowToStart+i)] = productData[i][7]
        sheet["n"+str(numRowToStart+i)] = productData[i][8]
        sheet["o"+str(numRowToStart+i)] = productData[i][9]
        sheet["p"+str(numRowToStart+i)] = float(productData[i][9]) + float(productData[i][7])
        sheet["q"+str(numRowToStart+i)] = productData[i][10]
        for col in range(1,18):
            sheet.cell(row = numRowToStart+i,column = col).border = thin_border
    wb.save("records/monthwise_record/"+str(date.year())+".xlsx")

def saveProductWise():
    numOfProducts = 0
    
    for c in range(14,-1,-1):
        if el_productnames[c].text() == "":
            numOfProducts += 1
        else:
            numOfProducts = 15 - numOfProducts
            break

    date = x.date_edit.date()

    ven_name = x.vendor_name.text()
    ven_gst = x.vedor_gst_no.text()
    invoice_number = x.invoice_number.text()

    i = -1
    for q in range(numOfProducts):
        if el_productnames[q].text()=="":
            continue
        i+=1
        wb = load_workbook("records/productwise_record/"+productData[i][0].split("-")[0]+".xlsx")
        sheet = wb["Sheet1"]
        numRowToStart = sheet.max_row +1
        print(numRowToStart)
        sheet["a"+str(numRowToStart)] = numRowToStart-2
        sheet["c"+str(numRowToStart)] = str(date.day())+"/"+str(date.month())+ "/" + str(date.year())        
        sheet["d"+str(numRowToStart)] = invoice_number
        sheet["e"+str(numRowToStart)] = ven_gst
        sheet["f"+str(numRowToStart)] = ven_name
        sheet["g"+str(numRowToStart)] = productData[i][2]
        sheet["h"+str(numRowToStart)] = productData[i][3]
        vol = calcVolume(productData[i][2],productData[i][3])
        sheet["i"+str(numRowToStart)] = vol
        sheet["j"+str(numRowToStart)] = productData[i][4]
        sheet["k"+str(numRowToStart)] = productData[i][5]
        sheet["l"+str(numRowToStart)] = productData[i][6]
        sheet["m"+str(numRowToStart)] = productData[i][7]
        sheet["n"+str(numRowToStart)] = productData[i][8]
        sheet["o"+str(numRowToStart)] = productData[i][9]
        sheet["p"+str(numRowToStart)] = float(productData[i][9]) + float(productData[i][7])
        sheet["q"+str(numRowToStart)] = productData[i][10]
        for col in range(1,18):
            sheet.cell(row = numRowToStart,column = col).border = thin_border
        wb.save("records/productwise_record/"+productData[i][0].split("-")[0]+".xlsx")

def saveVendorWise():
    ven_name = x.vendor_name.text()
    ven_gst = x.vedor_gst_no.text()
    invoice_number = x.invoice_number.text()
    date = x.date_edit.date()   

    numOfProducts = 0
    for c in range(14,-1,-1):
        if el_productnames[c].text() == "":
            numOfProducts += 1
        else:
            numOfProducts = 15 - numOfProducts
            break

    if not os.path.isfile("records/vendorwise_record/"+ven_name+".xlsx"):
        copyfile("data/vendor-wise-record.xlsx","records/vendorwise_record/"+ven_name+".xlsx")

    wb = load_workbook("records/vendorwise_record/"+ven_name+".xlsx")
    sheet = wb['Sheet1'] 
    numRowToStart = sheet.max_row +1
    print(numRowToStart)

    i=-1
    for q in range(numOfProducts):
        if el_productnames[q].text()=="":
            continue
        i+=1
        sheet["a"+str(numRowToStart+i)] = numRowToStart+i-2
        sheet["c"+str(numRowToStart+i)] = str(date.day())+"/"+str(date.month())+ "/" + str(date.year())        
        sheet["d"+str(numRowToStart+i)] = invoice_number
        sheet["e"+str(numRowToStart+i)] = ven_gst
        sheet["f"+str(numRowToStart+i)] = productData[i][0].split("-")[0]
        sheet["g"+str(numRowToStart+i)] = productData[i][1]
        sheet["h"+str(numRowToStart+i)] = productData[i][2]
        sheet["i"+str(numRowToStart+i)] = productData[i][3]
        vol = calcVolume(productData[i][2],productData[i][3])
        sheet["j"+str(numRowToStart+i)] = vol
        sheet["k"+str(numRowToStart+i)] = productData[i][4]
        sheet["l"+str(numRowToStart+i)] = productData[i][5]
        sheet["m"+str(numRowToStart+i)] = productData[i][6]
        sheet["n"+str(numRowToStart+i)] = productData[i][7]
        sheet["o"+str(numRowToStart+i)] = productData[i][8]
        sheet["p"+str(numRowToStart+i)] = productData[i][9]
        sheet["q"+str(numRowToStart+i)] = float(productData[i][9]) + float(productData[i][7])
        sheet["r"+str(numRowToStart+i)] = productData[i][10]
        for col in range(1,18):
            sheet.cell(row = numRowToStart+i,column = col).border = thin_border
    wb.save("records/vendorwise_record/"+ven_name+".xlsx")

def calcVolume(uom,qty):
    qty = int(qty)
    uom = uom.split(" ")
    vol_unit = ""
    vol_no = ""
    if uom[1] == "Kg":
        vol_no = int(uom[0]) * qty
        vol_unit = "Kg"
    if uom[1] == "L":
        vol_no = int(uom[0]) * qty
        vol_unit = "L"
    if uom[1] == "units":
        vol_no = int(uom[0]) * qty
        vol_unit = "units"
    if uom[1] == "gm":
        vol_no = int(uom[0]) * qty/1000
        vol_unit = "Kg"
    if uom[1] == "ml":
        vol_no = int(uom[0]) * qty/1000
        vol_unit = "L"

    return str(vol_no) + " " + vol_unit